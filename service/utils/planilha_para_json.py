"""
Converte planilhas Excel (.xlsx) da pasta docs/ em arquivos JSON.

Cada aba da planilha gera um JSON separado com o nome da aba.
Os JSONs são salvos na pasta resources/.

Uso:
    python planilha_para_json.py
"""

import glob
import json
import os
import sys
from typing import Dict, List

from openpyxl import load_workbook

DOCS_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "docs")
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "resources")


def listar_planilhas(pasta: str) -> List[str]:
    """Retorna caminhos absolutos de todos os .xlsx na pasta."""
    padrao = os.path.join(os.path.abspath(pasta), "**", "*.xlsx")
    return sorted(glob.glob(padrao, recursive=True))


def aba_para_lista(ws) -> List[Dict]:
    """Converte uma aba do workbook em lista de dicts (primeira linha = cabeçalho)."""
    linhas = list(ws.iter_rows(values_only=True))
    if len(linhas) < 2:
        return []

    cabecalho_raw = linhas[0]
    cabecalho = [str(c).strip() if c is not None else f"coluna_{i}" for i, c in enumerate(cabecalho_raw)]

    registros = []
    for linha in linhas[1:]:
        registro = {}
        for col, valor in zip(cabecalho, linha):
            # Converte para string; None → null no JSON
            registro[col] = str(valor) if valor is not None else None
        registros.append(registro)

    return registros


def converter_planilha(caminho: str):
    """Abre a planilha e salva um JSON por aba em resources/."""
    wb = load_workbook(caminho, read_only=True, data_only=True)
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    for nome_aba in wb.sheetnames:
        ws = wb[nome_aba]
        dados = aba_para_lista(ws)

        nome_arquivo = f"{nome_aba}.json"
        caminho_saida = os.path.join(OUTPUT_DIR, nome_arquivo)

        with open(caminho_saida, "w", encoding="utf-8") as f:
            json.dump(dados, f, indent=2, ensure_ascii=False)

        print(f"  ✔ {nome_aba} → {caminho_saida}  ({len(dados)} registro(s))")

    wb.close()


def main():
    planilhas = listar_planilhas(DOCS_DIR)

    if not planilhas:
        print(f"Nenhuma planilha .xlsx encontrada em {os.path.abspath(DOCS_DIR)}/")
        sys.exit(1)

    print("Planilhas encontradas:\n")
    for i, p in enumerate(planilhas, start=1):
        print(f"  [{i}] {os.path.basename(p)}")

    print()
    escolha = input("Digite o número da planilha desejada: ").strip()

    try:
        indice = int(escolha) - 1
        if indice < 0 or indice >= len(planilhas):
            raise ValueError
    except ValueError:
        print("Opção inválida.")
        sys.exit(1)

    planilha = planilhas[indice]
    print(f"\nConvertendo '{os.path.basename(planilha)}'...\n")
    converter_planilha(planilha)
    print("\nConversão concluída!")


if __name__ == "__main__":
    main()
